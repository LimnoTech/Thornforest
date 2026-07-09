import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from _helpers import save_workbook


def test_save_workbook_writes_one_sheet_per_entry_in_order(tmp_path):
    import openpyxl

    sheets = {
        "first": pd.DataFrame({"a": [1, 2]}),
        "second": pd.DataFrame({"b": [3, 4]}),
    }
    out = tmp_path / "book.xlsx"
    assert save_workbook(sheets, out) is None  # side-effect helper returns nothing

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["first", "second"]


def test_save_workbook_freezes_panes_and_sets_autofilter(tmp_path):
    import openpyxl

    sheets = {"data": pd.DataFrame({"a": [1, 2], "b": [3, 4]})}
    out = tmp_path / "book.xlsx"
    save_workbook(sheets, out)

    ws = openpyxl.load_workbook(out)["data"]
    assert ws.freeze_panes == "B2"
    assert ws.auto_filter.ref == "A1:B3"


def test_save_workbook_converts_geodataframe_geometry_to_wkt(tmp_path):
    import geopandas as gpd
    import openpyxl
    from shapely.geometry import Point

    gdf = gpd.GeoDataFrame(
        {"id": ["a", "b"], "geometry": [Point(0, 0), Point(1, 1)]}, crs="EPSG:4326"
    )
    out = tmp_path / "book.xlsx"
    save_workbook({"points": gdf}, out)

    ws = openpyxl.load_workbook(out)["points"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "geometry")
    assert rows[1] == ("a", "POINT (0 0)")
    assert rows[2] == ("b", "POINT (1 1)")


def test_save_workbook_truncates_and_deduplicates_long_sheet_names(tmp_path):
    import openpyxl

    long_name = "usgs_monitoring_locations_parameters"  # 37 chars, exceeds Excel's 31-char limit
    sheets = {
        long_name: pd.DataFrame({"a": [1]}),
        long_name + "_2": pd.DataFrame({"a": [2]}),  # truncates to the same 31-char prefix
    }
    out = tmp_path / "book.xlsx"
    save_workbook(sheets, out)

    names = openpyxl.load_workbook(out).sheetnames
    assert len(names) == 2
    assert len(set(names)) == 2  # unique
    assert all(len(n) <= 31 for n in names)
